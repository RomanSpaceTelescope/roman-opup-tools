#!/usr/bin/env python3
"""
Astronomical Sky Plotter Generator
Generates a standalone HTML file for plotting astronomical targets on an Aitoff projection.
Author: maxime.j.rizzo@nasa.gov
AI-assisted (Claude 4.6 Sonnet & Opus)
"""

import csv
import json
import os
import argparse

def embed_csv_as_js_array(csv_filename):
    """
    Read CSV file and convert to JavaScript array for embedding
    """    
    stars = []
    with open(csv_filename, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            # Only include non-empty values
            clean_row = {k: v for k, v in row.items() if v and v.strip()}
            stars.append(clean_row)
    
    return json.dumps(stars)

def embed_xlsx_as_js_array(xlsx_filename, sheet_name=None):
    """
    Read XLSX/XLS file and convert to JavaScript array for embedding.
    Requires openpyxl (for .xlsx) or xlrd (for .xls).
    
    Parameters
    ----------
    xlsx_filename : str
        Path to the Excel file.
    sheet_name : str or None
        Sheet to read. If None, reads the first/active sheet.
    
    Returns
    -------
    str
        JSON string of list-of-dicts (same format as embed_csv_as_js_array).
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "openpyxl is required to read .xlsx files. "
            "Install it with: pip install openpyxl"
        )

    wb = openpyxl.load_workbook(xlsx_filename, read_only=True, data_only=True)
    
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            available = ', '.join(wb.sheetnames)
            raise ValueError(
                f"Sheet '{sheet_name}' not found in '{xlsx_filename}'. "
                f"Available sheets: {available}"
            )
        ws = wb[sheet_name]
    else:
        ws = wb.active
    
    rows_iter = ws.iter_rows(values_only=True)
    
    # First row = header
    header = [str(h) if h is not None else f'_col{i}' for i, h in enumerate(next(rows_iter))]
    
    records = []
    for row in rows_iter:
        row_dict = {}
        for col_name, value in zip(header, row):
            if value is not None:
                # Convert everything to string to match CSV behavior
                row_dict[col_name] = str(value).strip()
        if row_dict:  # skip completely empty rows
            records.append(row_dict)
    
    wb.close()
    return json.dumps(records)


def embed_file_as_js_array(filename, sheet_name=None):
    """
    Unified reader: dispatches to CSV or XLSX reader based on file extension.
    
    Parameters
    ----------
    filename : str
        Path to .csv, .xlsx, or .xls file.
    sheet_name : str or None
        For Excel files, which sheet to read. Ignored for CSV.
    
    Returns
    -------
    str
        JSON string of list-of-dicts.
    """
    ext = os.path.splitext(filename)[1].lower()
    
    if ext == '.csv':
        return embed_csv_as_js_array(filename)
    elif ext in ('.xlsx', '.xls'):
        return embed_xlsx_as_js_array(filename, sheet_name=sheet_name)
    else:
        raise ValueError(
            f"Unsupported file type '{ext}' for '{filename}'. "
            f"Supported types: .csv, .xlsx, .xls"
        )
    

def get_sun_position(date=None, exclusion_radius=54, n_circle_pts=361):
    """
    Compute the Sun's position, anti-Sun position, and exclusion zone circles.
    
    Parameters
    ----------
    date : datetime or None
    exclusion_radius : float
        Angular radius of exclusion zone in degrees (default 54 for Roman).
    n_circle_pts : int
        Number of points to generate around each circle.
    
    Returns
    -------
    dict with keys: ra, dec, l, b, date_str,
                    anti_ra, anti_dec, anti_l, anti_b,
                    sun_circle (list of {ra, dec, l, b}),
                    anti_sun_circle (list of {ra, dec, l, b})
    """
    import math
    from datetime import datetime, timezone

    if date is None:
        date = datetime.now(timezone.utc)

    # Days since J2000.0
    n = (date - datetime(2000, 1, 1, 12, 0, 0, tzinfo=timezone.utc)).total_seconds() / 86400.0

    # Mean longitude and mean anomaly (degrees)
    L = (280.460 + 0.9856474 * n) % 360
    g = (357.528 + 0.9856003 * n) % 360
    g_rad = math.radians(g)

    # Ecliptic longitude
    lam = (L + 1.915 * math.sin(g_rad) + 0.020 * math.sin(2 * g_rad)) % 360
    lam_rad = math.radians(lam)

    # Obliquity
    eps = 23.439 - 0.0000004 * n
    eps_rad = math.radians(eps)

    # Sun equatorial coordinates
    ra = math.degrees(math.atan2(math.cos(eps_rad) * math.sin(lam_rad), math.cos(lam_rad)))
    if ra < 0:
        ra += 360
    dec = math.degrees(math.asin(math.sin(eps_rad) * math.sin(lam_rad)))

    # Anti-Sun (180 degrees opposite)
    anti_ra = (ra + 180) % 360
    anti_dec = -dec

    def eq_to_gal(ra_deg, dec_deg):
        """Equatorial to galactic."""
        ra_r = math.radians(ra_deg)
        dec_r = math.radians(dec_deg)
        x_eq = math.cos(dec_r) * math.cos(ra_r)
        y_eq = math.cos(dec_r) * math.sin(ra_r)
        z_eq = math.sin(dec_r)
        T = [
            [-0.054875539726, -0.873437108010, -0.483834985808],
            [+0.494109453312, -0.444829589425, +0.746982251810],
            [-0.867666135858, -0.198076386122, +0.455983795705]
        ]
        x_g = T[0][0]*x_eq + T[0][1]*y_eq + T[0][2]*z_eq
        y_g = T[1][0]*x_eq + T[1][1]*y_eq + T[1][2]*z_eq
        z_g = T[2][0]*x_eq + T[2][1]*y_eq + T[2][2]*z_eq
        b = math.degrees(math.asin(max(-1, min(1, z_g))))
        l = math.degrees(math.atan2(y_g, x_g))
        if l < 0:
            l += 360
        return l, b

    def circle_around(center_ra, center_dec, radius_deg, n_pts):
        """
        Generate points on a circle of angular radius around a center point.
        Uses rotation: place circle at pole, then rotate pole to center.
        """
        pts = []
        cr = math.radians(center_ra)
        cd = math.radians(center_dec)
        r = math.radians(radius_deg)

        for i in range(n_pts):
            az = -2 * math.pi * i / (n_pts - 1)  # azimuth around circle

            # Point at angular distance r from north pole, azimuth az
            x = math.sin(r) * math.cos(az)
            y = math.sin(r) * math.sin(az)
            z = math.cos(r)

            # Rotate from pole (0,0,1) to center (cr, cd):
            # First rotate by (90° - dec) around Y-axis, then by ra around Z-axis
            # Ry(90-dec): rotate so pole moves to declination cd
            cos_t = math.sin(cd)  # cos(90-dec) = sin(dec)
            sin_t = math.cos(cd)  # sin(90-dec) = cos(dec)
            x1 = cos_t * x + sin_t * z
            y1 = y
            z1 = -sin_t * x + cos_t * z

            # Rz(ra): rotate around Z-axis by RA
            x2 = math.cos(cr) * x1 - math.sin(cr) * y1
            y2 = math.sin(cr) * x1 + math.cos(cr) * y1
            z2 = z1

            # Back to spherical
            pt_dec = math.degrees(math.asin(max(-1, min(1, z2))))
            pt_ra = math.degrees(math.atan2(y2, x2))
            if pt_ra < 0:
                pt_ra += 360

            pt_l, pt_b = eq_to_gal(pt_ra, pt_dec)
            pts.append({'ra': pt_ra, 'dec': pt_dec, 'l': pt_l, 'b': pt_b})

        return pts

    # Compute galactic coords for Sun and anti-Sun
    sun_l, sun_b = eq_to_gal(ra, dec)
    anti_l, anti_b = eq_to_gal(anti_ra, anti_dec)

    # Generate exclusion zone circles
    sun_circle = circle_around(ra, dec, exclusion_radius, n_circle_pts)
    anti_sun_circle = circle_around(anti_ra, anti_dec, exclusion_radius, n_circle_pts)

    date_str = date.strftime('%Y-%m-%d')

    return {
        'ra': ra, 'dec': dec, 'l': sun_l, 'b': sun_b,
        'anti_ra': anti_ra, 'anti_dec': anti_dec, 'anti_l': anti_l, 'anti_b': anti_b,
        'sun_circle': sun_circle,
        'anti_sun_circle': anti_sun_circle,
        'date_str': date_str,
        'exclusion_radius': exclusion_radius
    }

def parse_start_date(start_val):
    """
    Parse a 'Start' field value into a datetime.
    Handles spacecraft DOY format (2026-275-17:22:08 TAI), ISO, US, MJD.
    
    Returns datetime (UTC) or None.
    """
    import re
    from datetime import datetime, timezone, timedelta

    if not start_val:
        return None
    s = start_val.strip()

    # YYYY-DDD or YYYY-DDD-HH:MM:SS with optional timescale suffix
    doy_match = re.match(
        r'^(\d{4})-(\d{1,3})(?:-(\d{2}:\d{2}:\d{2}))?\s*(?:TAI|UTC|TDB|TT)?$', s, re.I
    )
    if doy_match:
        year = int(doy_match.group(1))
        doy = int(doy_match.group(2))
        if 1 <= doy <= 366:
            d = datetime(year, 1, 1, tzinfo=timezone.utc) + timedelta(days=doy - 1)
            if doy_match.group(3):
                hh, mm, ss = map(int, doy_match.group(3).split(':'))
                d = d.replace(hour=hh, minute=mm, second=ss)
            return d

    # Common calendar formats
    for fmt in ('%Y-%m-%d', '%Y-%m-%dT%H:%M:%S', '%Y-%m-%dT%H:%M:%SZ',
                '%m/%d/%Y', '%d/%m/%Y', '%b %d, %Y', '%Y-%m-%d %H:%M:%S'):
        try:
            return datetime.strptime(s, fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            continue

    # MJD
    try:
        mjd = float(s)
        if 40000 < mjd < 100000:
            return datetime(1858, 11, 17, tzinfo=timezone.utc) + timedelta(days=mjd)
    except (ValueError, TypeError):
        pass

    return None


def generate_html(preloaded_datasets=None, sun_position=None):

    star_catalog_json = embed_csv_as_js_array('Constellation_Stars_nolatex.csv')

    # Build preloaded datasets JS
    if preloaded_datasets:
        entries = []
        for ds in preloaded_datasets:
            entries.append(
                f'{{ fileName: {json.dumps(ds["fileName"])}, rawData: {ds["data_json"]} }}'
            )
        preloaded_js = "const PRELOADED_DATASETS = [" + ",\n            ".join(entries) + "];"
    else:
        preloaded_js = "const PRELOADED_DATASETS = [];"
    
    # Build Sun position JS
    # Build Sun position JS (including exclusion zone circles)
    if sun_position:
        sun_circle_eq = [{'ra': p['ra'], 'dec': p['dec']} for p in sun_position['sun_circle']]
        sun_circle_gal = [{'l': p['l'], 'b': p['b']} for p in sun_position['sun_circle']]
        anti_circle_eq = [{'ra': p['ra'], 'dec': p['dec']} for p in sun_position['anti_sun_circle']]
        anti_circle_gal = [{'l': p['l'], 'b': p['b']} for p in sun_position['anti_sun_circle']]

        sun_js = (
            f"const SUN_POSITION = {{"
            f" ra: {sun_position['ra']}, dec: {sun_position['dec']},"
            f" l: {sun_position['l']}, b: {sun_position['b']},"
            f" anti_ra: {sun_position['anti_ra']}, anti_dec: {sun_position['anti_dec']},"
            f" anti_l: {sun_position['anti_l']}, anti_b: {sun_position['anti_b']},"
            f" dateStr: {json.dumps(sun_position['date_str'])},"
            f" exclusionRadius: {sun_position['exclusion_radius']},"
            f" sunCircleEq: {json.dumps(sun_circle_eq)},"
            f" sunCircleGal: {json.dumps(sun_circle_gal)},"
            f" antiCircleEq: {json.dumps(anti_circle_eq)},"
            f" antiCircleGal: {json.dumps(anti_circle_gal)}"
            f" }};"
        )
    else:
        sun_js = "const SUN_POSITION = null;"

    html_content = r"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Roman Visit Plotter</title>
    <script src="https://unpkg.com/react@18/umd/react.production.min.js"></script>
    <script src="https://unpkg.com/react-dom@18/umd/react-dom.production.min.js"></script>
    <script src="https://unpkg.com/@babel/standalone/babel.min.js"></script>
    <script src="https://cdn.plot.ly/plotly-2.27.0.min.js"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/PapaParse/5.4.1/papaparse.min.js"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/xlsx/0.18.5/xlsx.full.min.js"></script>
    <script src="https://cdn.tailwindcss.com"></script>
</head>
<body>
    <div id="root"></div>
    
    <script type="text/babel">
        const { useState, useEffect } = React;

        // ===== PRE-LOADED DATASETS (embedded at generation time) =====
        __PRELOADED_PLACEHOLDER__
        __SUN_PLACEHOLDER__

        function AstronomicalSkyPlotter() {
          const [datasets, setDatasets] = useState([]);
          const [plotKey, setPlotKey] = useState(0);
          const [coordSystem, setCoordSystem] = useState('equatorial'); // 'equatorial' or 'galactic'
          const [showStars, setShowStars] = React.useState(true); // NEW: toggle for stars

          // Color palette for different purposes
          const colorPalette = [
            '#FF6B6B', '#4ECDC4', '#45B7D1', '#FFA07A', '#98D8C8',
            '#F7DC6F', '#BB8FCE', '#85C1E2', '#F8B739', '#52B788'
          ];

          const STAR_CATALOG = """ 

    html_content += star_catalog_json

    html_content += r""";

          const getStarCatalog = () => {{
            return STAR_CATALOG;
          }};

          // Convert equatorial (RA, Dec) to galactic (l, b) coordinates
        function equatorialToGalactic(ra, dec) {
            // Convert input to radians
            const raRad = ra * Math.PI / 180;
            const decRad = dec * Math.PI / 180;
            
            // Convert equatorial to Cartesian coordinates
            const x_eq = Math.cos(decRad) * Math.cos(raRad);
            const y_eq = Math.cos(decRad) * Math.sin(raRad);
            const z_eq = Math.sin(decRad);
            
            // Rotation matrix from equatorial to galactic (J2000)
            // These values are from the IAU definition
            const T11 = -0.054875539726;
            const T12 = -0.873437108010;
            const T13 = -0.483834985808;
            const T21 = +0.494109453312;
            const T22 = -0.444829589425;
            const T23 = +0.746982251810;
            const T31 = -0.867666135858;
            const T32 = -0.198076386122;
            const T33 = +0.455983795705;
            
            // Apply rotation matrix
            const x_gal = T11 * x_eq + T12 * y_eq + T13 * z_eq;
            const y_gal = T21 * x_eq + T22 * y_eq + T23 * z_eq;
            const z_gal = T31 * x_eq + T32 * y_eq + T33 * z_eq;
            
            // Convert Cartesian galactic to spherical galactic
            const b = Math.asin(z_gal) * 180 / Math.PI;
            let l = Math.atan2(y_gal, x_gal) * 180 / Math.PI;
            
            // Normalize l to [0, 360)
            if (l < 0) l += 360;
            
            return { l: l, b: b };
        }

          // Generate ecliptic plane trace (in equatorial coordinates)
          const getEclipticTrace = () => {
            const eclipticObliquity = 23.43928; // degrees
            const points = [];
            
            for (let lambda = 0; lambda <= 360; lambda += 1) {
              const lambdaRad = lambda * Math.PI / 180;
              const oblRad = eclipticObliquity * Math.PI / 180;
              
              // Convert ecliptic to equatorial
              const raRad = Math.atan2(Math.sin(lambdaRad) * Math.cos(oblRad), Math.cos(lambdaRad));
              const decRad = Math.asin(Math.sin(lambdaRad) * Math.sin(oblRad));
              
              let ra = raRad * 180 / Math.PI;
              if (ra < 0) ra += 360;
              const dec = decRad * 180 / Math.PI;
              
              points.push({ ra, dec });
            }
            
            return points;
          };

          // Generate galactic plane trace (in equatorial coordinates)
          const getGalacticPlaneInEquatorial = () => {
            const points = [];
            
            // Generate points along galactic equator (b=0)
            for (let l = 0; l <= 360; l += 1) {
              const lRad = l * Math.PI / 180;
              const b = 0;
              const bRad = 0;
              
              // North Galactic Pole in equatorial coordinates (J2000)
              const alphaNGP = 192.85948 * Math.PI / 180;
              const deltaNGP = 27.12825 * Math.PI / 180;
              const lNCP = 122.93192 * Math.PI / 180;
              
              // Convert galactic to equatorial (inverse transformation)
              const sinDec = Math.cos(bRad) * Math.cos(deltaNGP) * Math.sin(lRad - lNCP) + 
                            Math.sin(bRad) * Math.sin(deltaNGP);
              const dec = Math.asin(sinDec) * 180 / Math.PI;
              
              const y = Math.cos(bRad) * Math.cos(lRad - lNCP);
              const x = Math.sin(bRad) * Math.cos(deltaNGP) - 
                       Math.cos(bRad) * Math.sin(deltaNGP) * Math.sin(lRad - lNCP);
              
              let ra = (alphaNGP + Math.atan2(y, x)) * 180 / Math.PI;
              if (ra < 0) ra += 360;
              if (ra >= 360) ra -= 360;
              
              points.push({ ra, dec });
            }
            
            return points;
          };

          // Generate continuous viewing zone lines at +/- 54 degrees ecliptic latitude
          const getContinuousViewingZone = (eclipticLat) => {
            const eclipticObliquity = 23.43928; // degrees
            const points = [];
            
            for (let lambda = 0; lambda <= 360; lambda += 1) {
              const lambdaRad = lambda * Math.PI / 180;
              const betaRad = eclipticLat * Math.PI / 180;
              const oblRad = eclipticObliquity * Math.PI / 180;
              
              // Convert ecliptic (lambda, beta) to equatorial (RA, Dec)
              const raRad = Math.atan2(
                Math.sin(lambdaRad) * Math.cos(oblRad) - Math.tan(betaRad) * Math.sin(oblRad),
                Math.cos(lambdaRad)
              );
              const decRad = Math.asin(
                Math.sin(betaRad) * Math.cos(oblRad) + Math.cos(betaRad) * Math.sin(oblRad) * Math.sin(lambdaRad)
              );
              
              let ra = raRad * 180 / Math.PI;
              if (ra < 0) ra += 360;
              const dec = decRad * 180 / Math.PI;
              
              points.push({ ra, dec });
            }
            
            return points;
          };

          const handleFileUpload = (event) => {
            const file = event.target.files[0];
            if (!file) return;

            const fileName = file.name;
            const fileExtension = fileName.split('.').pop().toLowerCase();

            if (fileExtension === 'csv') {
              // Parse CSV
              Papa.parse(file, {
                header: true,
                dynamicTyping: true,
                complete: (results) => {
                  processData(results.data, fileName);
                },
                error: (error) => {
                  alert(`Error parsing CSV: ${error.message}`);
                }
              });
            } else if (['xlsx', 'xls'].includes(fileExtension)) {
              // Parse Excel
              const reader = new FileReader();
              reader.onload = (e) => {
                try {
                  const data = new Uint8Array(e.target.result);
                  const workbook = XLSX.read(data, { type: 'array' });
                  const firstSheet = workbook.Sheets[workbook.SheetNames[0]];
                  const jsonData = XLSX.utils.sheet_to_json(firstSheet);
                  processData(jsonData, fileName);
                } catch (error) {
                  alert(`Error parsing Excel: ${error.message}`);
                }
              };
              reader.readAsArrayBuffer(file);
            } else {
              alert('Please upload a CSV or Excel file');
            }
          };

          const processData = (rawRows, fileName) => {
            // Filter for rows that have RA and DEC
            const validData = rawRows.filter(row => {
              const hasRA = row.RA !== undefined || row.ra !== undefined || row.Ra !== undefined;
              const hasDEC = row.DEC !== undefined || row.dec !== undefined || row.Dec !== undefined;
              return hasRA && hasDEC;
            });

            if (validData.length === 0) {
              alert(`No valid RA/DEC data found in ${fileName}`);
              return;
            }

            // Extract RA, DEC, and Intended_Purpose, plus all other columns
            const processedData = validData.map(row => {
              const ra = parseFloat(row.RA || row.ra || row.Ra);
              const dec = parseFloat(row.DEC || row.dec || row.Dec);
              const purpose = row.Intended_Purpose || row.intended_purpose || row.INTENDED_PURPOSE || 'Unknown';
              const name = row.Name || row.name || row.NAME || '';
              
              const allColumns = {};
              Object.keys(row).forEach(key => {
                if (['RA', 'ra', 'Ra', 'DEC', 'dec', 'Dec', 'Intended_Purpose', 'intended_purpose', 'INTENDED_PURPOSE'].includes(key)) {
                  return;
                }
                if (row[key] !== null && row[key] !== undefined && row[key] !== '') {
                  let columnName = key;
                  if (!key || key.trim() === '' || key.startsWith('_') || 
                      key.toLowerCase().includes('unnamed') || key.includes('__EMPTY')) {
                    columnName = 'Index';
                  }
                  allColumns[columnName] = row[key];
                }
              });
              
              return { ra, dec, purpose, name, allColumns, rawData: row };
            });

            const newDataset = {
              fileName: fileName,
              data: processedData.filter(d => !isNaN(d.ra) && !isNaN(d.dec))
            };

            setDatasets(prev => [...prev, newDataset]);
            setPlotKey(prev => prev + 1);
          };


          const removeDataset = (index) => {
            setDatasets(prev => prev.filter((_, i) => i !== index));
            setPlotKey(prev => prev + 1);
          };

          const clearAll = () => {
            setDatasets([]);
            setPlotKey(prev => prev + 1);
          };

          // ===== NEW: Load pre-embedded datasets on mount =====
          useEffect(() => {
            if (typeof PRELOADED_DATASETS !== 'undefined' && PRELOADED_DATASETS.length > 0) {
              console.log(`Loading ${PRELOADED_DATASETS.length} pre-loaded dataset(s)...`);
              PRELOADED_DATASETS.forEach(ds => {
                processData(ds.rawData, ds.fileName + ' (pre-loaded)');
              });
            }
          }, []);  // empty dependency array = run once on mount


          useEffect(() => {
            if (datasets.length === 0) return;

            // Combine all datasets and get unique purposes
            const allData = datasets.flatMap(ds => ds.data);
            const purposes = [...new Set(allData.map(d => d.purpose))];
            
            // Create color mapping
            const purposeColors = {};
            purposes.forEach((purpose, idx) => {
              purposeColors[purpose] = colorPalette[idx % colorPalette.length];
            });

            // Map optical element to marker symbol
            const getMarkerSymbol = (row) => {
              const elem = (row.WFI_Optical_Element || row.wfi_optical_element 
                         || row.WFI_OPTICAL_ELEMENT || '').toUpperCase().trim();
              if (elem.includes('GRISM')) return 'triangle-up';
              if (elem.includes('PRISM')) return 'diamond';
              return 'circle';
            };

            const traces = purposes.map(purpose => {
              const purposeData = allData.filter(d => d.purpose === purpose);

              let lon, lat, coordLabels;

              if (coordSystem === 'equatorial') {
                lon = purposeData.map(d => d.ra > 180 ? d.ra - 360 : d.ra);
                lat = purposeData.map(d => d.dec);
                coordLabels = purposeData.map(d => `<br>RA: ${d.ra.toFixed(4)}°<br>Dec: ${d.dec.toFixed(4)}°`);
              } else {
                const galactic = purposeData.map(d => equatorialToGalactic(d.ra, d.dec));
                lon = galactic.map(g => {
                  let l_centered = g.l;
                  if (l_centered > 180) l_centered -= 360;
                  return l_centered;
                });
                lat = galactic.map(g => g.b);
                coordLabels = galactic.map(g => `<br>l: ${g.l.toFixed(4)}°<br>b: ${g.b.toFixed(4)}°`);
              }

              // Build per-point symbol array from optical element
              const symbols = purposeData.map(d => getMarkerSymbol(d.rawData || {}));

              return {
                type: 'scattergeo',
                mode: 'markers',
                lon: lon,
                lat: lat,
                text: purposeData.map((d, idx) => {
                  let hoverText = `<b>${purpose}</b>`;
                  if (d.name) hoverText += `<br>Name: ${d.name}`;
                  hoverText += coordLabels[idx];
                  if (d.allColumns && Object.keys(d.allColumns).length > 0) {
                    hoverText += '<br><br><b>Additional Data:</b>';
                    Object.keys(d.allColumns).forEach(key => {
                      const value = d.allColumns[key];
                      if (value !== null && value !== undefined && value !== '') {
                        hoverText += `<br>${key}: ${value}`;
                      }
                    });
                  }
                  return hoverText;
                }),
                hoverinfo: 'text',
                name: purpose,
                marker: {
                  size: 8,
                  color: purposeColors[purpose],
                  opacity: 0.7,
                  symbol: symbols,
                  line: {
                    color: 'white',
                    width: 0.5
                  }
                },
                showlegend: true
              };
            });

            // =============================================
            // TRACE GROUP 1: Star catalog (background layer)
            // =============================================
            
            // Add stars from catalog (if enabled)
            if (showStars) {
              const starCatalog = getStarCatalog();
              
              if (starCatalog && starCatalog.length > 0) {
                const starRAs = starCatalog.map(s => parseFloat(s['RA [deg]']));
                const starDecs = starCatalog.map(s => parseFloat(s['Dec [deg]']));
                const starVMags = starCatalog.map(s => parseFloat(s['V App Mag']) || 5);
                
                const starHoverText = starCatalog.map(star => {
                  let lines = [];
                  
                  if (star['Common Name']) lines.push(`<b>${star['Common Name']}</b>`);
                  if (star['I']) lines.push(`${star['I']}`);
                  if (star['Constellation']) lines.push(`${star['Constellation']}`);
                  
                  lines.push(`RA: ${parseFloat(star['RA [deg]']).toFixed(4)}°`);
                  lines.push(`Dec: ${parseFloat(star['Dec [deg]']).toFixed(4)}°`);
                  
                  if (star['V App Mag']) lines.push(`V mag: ${star['V App Mag']}`);
                  if (star['B-V']) lines.push(`B-V: ${star['B-V']}`);
                  if (star['Spec Type']) lines.push(`Type: ${star['Spec Type']}`);
                  if (star['Dist [pc]']) lines.push(`Dist: ${parseFloat(star['Dist [pc]']).toFixed(2)} pc`);
                  if (star['Mass [solar]']) lines.push(`Mass: ${star['Mass [solar]']} M☉`);
                  if (star['Rad [solar]']) lines.push(`Radius: ${star['Rad [solar]']} R☉`);
                  if (star['Lum [solar]']) lines.push(`Lum: ${star['Lum [solar]']} L☉`);
                  if (star['Temp [K]']) lines.push(`Temp: ${star['Temp [K]']} K`);
                  if (star['Prop M. RA [mas/yr]']) lines.push(`PM RA: ${star['Prop M. RA [mas/yr]']} mas/yr`);
                  if (star['Prop M. Dec [mas/yr]']) lines.push(`PM Dec: ${star['Prop M. Dec [mas/yr]']} mas/yr`);
                  if (star['Radial Velocity [km/s]']) lines.push(`RV: ${star['Radial Velocity [km/s]']} km/s`);
                  
                  return lines.join('<br>');
                });
                
                if (coordSystem === 'equatorial') {
                  traces.push({
                    type: 'scattergeo',
                    mode: 'markers',
                    lon: starRAs.map(ra => ra > 180 ? ra - 360 : ra),
                    lat: starDecs,
                    name: 'Catalog Stars',
                    text: starHoverText,
                    hoverinfo: 'text',
                    marker: {
                      size: starVMags.map(mag => Math.max(2, 10 - mag * 1.5)),
                      color: '#FFFFE0',
                      opacity: starVMags.map(mag => Math.min(0.7, 2.5 / (mag + 2))),
                      symbol: 'star',
                      line: {
                        color: '#FFFFFF',
                        width: 0.3
                      }
                    },
                    showlegend: true
                  });
                } else {
                  const starGalactic = starRAs.map((ra, idx) => equatorialToGalactic(ra, starDecs[idx]));
                  
                  traces.push({
                    type: 'scattergeo',
                    mode: 'markers',
                    lon: starGalactic.map(g => g.l > 180 ? g.l - 360 : g.l),
                    lat: starGalactic.map(g => g.b),
                    name: 'Catalog Stars',
                    text: starHoverText,
                    hoverinfo: 'text',
                    marker: {
                      size: starVMags.map(mag => Math.max(2, 10 - mag * 1.5)),
                      color: '#FFFFE0',
                      opacity: starVMags.map(mag => Math.min(0.7, 2.5 / (mag + 2))),
                      symbol: 'star',
                      line: {
                        color: '#FFFFFF',
                        width: 0.3
                      }
                    },
                    showlegend: true
                  });
                }
              }
            }

            // =============================================
            // TRACE GROUP 2: SUN EXCLUSION ZONES & MARKERS
            // =============================================

            // Add Sun, anti-Sun, and exclusion zones (all pre-computed in Python)
            if (SUN_POSITION) {
              const R = SUN_POSITION.exclusionRadius;

              // --- Sun exclusion zone circle ---
              const sunCircle = coordSystem === 'equatorial'
                ? SUN_POSITION.sunCircleEq
                : SUN_POSITION.sunCircleGal;
              const sunCircleLon = coordSystem === 'equatorial'
                ? sunCircle.map(p => p.ra > 180 ? p.ra - 360 : p.ra)
                : sunCircle.map(p => p.l > 180 ? p.l - 360 : p.l);
              const sunCircleLat = coordSystem === 'equatorial'
                ? sunCircle.map(p => p.dec)
                : sunCircle.map(p => p.b);

              traces.push({
                type: 'scattergeo',
                mode: 'lines',
                lon: sunCircleLon,
                lat: sunCircleLat,
                name: `Sun exclusion (${R}°)`,
                line: {
                  color: '#FF4500',
                  width: 2,
                  dash: 'dash'
                },
                fill: 'toself',
                fillcolor: 'rgba(180, 200, 255, 0.2)',
                hoverinfo: 'name',
                showlegend: true
              });

              // --- Anti-Sun exclusion zone circle ---
              const antiCircle = coordSystem === 'equatorial'
                ? SUN_POSITION.antiCircleEq
                : SUN_POSITION.antiCircleGal;
              const antiCircleLon = coordSystem === 'equatorial'
                ? antiCircle.map(p => p.ra > 180 ? p.ra - 360 : p.ra)
                : antiCircle.map(p => p.l > 180 ? p.l - 360 : p.l);
              const antiCircleLat = coordSystem === 'equatorial'
                ? antiCircle.map(p => p.dec)
                : antiCircle.map(p => p.b);

              traces.push({
                type: 'scattergeo',
                mode: 'lines',
                lon: antiCircleLon,
                lat: antiCircleLat,
                name: `Anti-Sun exclusion (${R}°)`,
                line: {
                  color: '#4169E1',
                  width: 2,
                  dash: 'dash'
                },
                fill: 'toself',
                fillcolor: 'rgba(180, 200, 255, 0.2)',
                hoverinfo: 'name',
                showlegend: true
              });

              // --- Sun marker ---
              let sunLon, sunLat, sunHoverCoords;
              if (coordSystem === 'equatorial') {
                sunLon = SUN_POSITION.ra > 180 ? SUN_POSITION.ra - 360 : SUN_POSITION.ra;
                sunLat = SUN_POSITION.dec;
                sunHoverCoords = `RA: ${SUN_POSITION.ra.toFixed(4)}°<br>Dec: ${SUN_POSITION.dec.toFixed(4)}°`;
              } else {
                sunLon = SUN_POSITION.l > 180 ? SUN_POSITION.l - 360 : SUN_POSITION.l;
                sunLat = SUN_POSITION.b;
                sunHoverCoords = `l: ${SUN_POSITION.l.toFixed(4)}°<br>b: ${SUN_POSITION.b.toFixed(4)}°`
                               + `<br><br>RA: ${SUN_POSITION.ra.toFixed(4)}°<br>Dec: ${SUN_POSITION.dec.toFixed(4)}°`;
              }

              traces.push({
                type: 'scattergeo',
                mode: 'markers+text',
                lon: [sunLon],
                lat: [sunLat],
                name: `Sun (${SUN_POSITION.dateStr})`,
                text: ['☀'],
                textposition: 'top center',
                textfont: { size: 18, color: '#FFD700' },
                marker: {
                  size: 14,
                  color: '#FFD700',
                  symbol: 'circle',
                  line: { color: '#FF8C00', width: 2 }
                },
                hoverinfo: 'text',
                hovertext: [`<b>Sun</b><br>Date: ${SUN_POSITION.dateStr}<br>${sunHoverCoords}`],
                showlegend: true
              });

              // --- Anti-Sun marker ---
              let antiLon, antiLat, antiHoverCoords;
              if (coordSystem === 'equatorial') {
                antiLon = SUN_POSITION.anti_ra > 180 ? SUN_POSITION.anti_ra - 360 : SUN_POSITION.anti_ra;
                antiLat = SUN_POSITION.anti_dec;
                antiHoverCoords = `RA: ${SUN_POSITION.anti_ra.toFixed(4)}°<br>Dec: ${SUN_POSITION.anti_dec.toFixed(4)}°`;
              } else {
                antiLon = SUN_POSITION.anti_l > 180 ? SUN_POSITION.anti_l - 360 : SUN_POSITION.anti_l;
                antiLat = SUN_POSITION.anti_b;
                antiHoverCoords = `l: ${SUN_POSITION.anti_l.toFixed(4)}°<br>b: ${SUN_POSITION.anti_b.toFixed(4)}°`
                                + `<br><br>RA: ${SUN_POSITION.anti_ra.toFixed(4)}°<br>Dec: ${SUN_POSITION.anti_dec.toFixed(4)}°`;
              }

              traces.push({
                type: 'scattergeo',
                mode: 'markers',
                lon: [antiLon],
                lat: [antiLat],
                name: `Anti-Sun (${SUN_POSITION.dateStr})`,
                marker: {
                  size: 12,
                  color: '#4169E1',
                  symbol: 'circle',
                  line: { color: '#1E3A8A', width: 2 }
                },
                hoverinfo: 'text',
                hovertext: [`<b>Anti-Sun</b><br>Date: ${SUN_POSITION.dateStr}<br>${antiHoverCoords}`],
                showlegend: true
              });
            }

            // Add ecliptic plane in equatorial mode
            if (coordSystem === 'equatorial') {
              const eclipticPoints = getEclipticTrace();
              traces.push({
                type: 'scattergeo',
                mode: 'lines',
                lon: eclipticPoints.map(p => p.ra > 180 ? p.ra - 360 : p.ra),
                lat: eclipticPoints.map(p => p.dec),
                name: 'Ecliptic Plane',
                line: {
                  color: '#FFD700',
                  width: 2,
                  dash: 'dot'
                },
                hoverinfo: 'name',
                showlegend: true
              });

              // Add continuous viewing zones at +/- 54 degrees ecliptic latitude
              const cvzPlus = getContinuousViewingZone(54);
              traces.push({
                type: 'scattergeo',
                mode: 'lines',
                lon: cvzPlus.map(p => p.ra > 180 ? p.ra - 360 : p.ra),
                lat: cvzPlus.map(p => p.dec),
                name: 'CVZ +54° (Ecliptic)',
                line: {
                  color: '#00CED1',
                  width: 1.5,
                  dash: 'dash'
                },
                hoverinfo: 'name',
                showlegend: true
              });

              const cvzMinus = getContinuousViewingZone(-54);
              traces.push({
                type: 'scattergeo',
                mode: 'lines',
                lon: cvzMinus.map(p => p.ra > 180 ? p.ra - 360 : p.ra),
                lat: cvzMinus.map(p => p.dec),
                name: 'CVZ -54° (Ecliptic)',
                line: {
                  color: '#00CED1',
                  width: 1.5,
                  dash: 'dash'
                },
                hoverinfo: 'name',
                showlegend: true
              });

              // Add galactic plane in equatorial coordinates
              const galacticPlanePoints = getGalacticPlaneInEquatorial();
              traces.push({
                type: 'scattergeo',
                mode: 'lines',
                lon: galacticPlanePoints.map(p => p.ra > 180 ? p.ra - 360 : p.ra),
                lat: galacticPlanePoints.map(p => p.dec),
                name: 'Galactic Plane',
                line: {
                  color: '#FF69B4',
                  width: 2,
                  dash: 'dash'
                },
                hoverinfo: 'name',
                showlegend: true
              });
            } else {
              // In galactic mode, just show the galactic equator (b=0)
              const galacticEquator = [];
              for (let l = 0; l <= 360; l += 1) {
                galacticEquator.push({ l, b: 0 });
              }
              traces.push({
                type: 'scattergeo',
                mode: 'lines',
                lon: galacticEquator.map(p => p.l > 180 ? p.l - 360 : p.l),
                lat: galacticEquator.map(p => p.b),
                name: 'Galactic Plane',
                line: {
                  color: '#FF69B4',
                  width: 3
                },
                hoverinfo: 'name',
                showlegend: true
              });

              // Add ecliptic in galactic coordinates
              const eclipticInGalactic = getEclipticTrace().map(p => equatorialToGalactic(p.ra, p.dec));
              traces.push({
                type: 'scattergeo',
                mode: 'lines',
                lon: eclipticInGalactic.map(p => p.l > 180 ? p.l - 360 : p.l),
                lat: eclipticInGalactic.map(p => p.b),
                name: 'Ecliptic Plane',
                line: {
                  color: '#FFD700',
                  width: 2,
                  dash: 'dot'
                },
                hoverinfo: 'name',
                showlegend: true
              });

              // Add continuous viewing zones in galactic coordinates
              const cvzPlusInGalactic = getContinuousViewingZone(54).map(p => equatorialToGalactic(p.ra, p.dec));
              traces.push({
                type: 'scattergeo',
                mode: 'lines',
                lon: cvzPlusInGalactic.map(p => p.l > 180 ? p.l - 360 : p.l),
                lat: cvzPlusInGalactic.map(p => p.b),
                name: 'CVZ +54° (Ecliptic)',
                line: {
                  color: '#00CED1',
                  width: 1.5,
                  dash: 'dash'
                },
                hoverinfo: 'name',
                showlegend: true
              });

              const cvzMinusInGalactic = getContinuousViewingZone(-54).map(p => equatorialToGalactic(p.ra, p.dec));
              traces.push({
                type: 'scattergeo',
                mode: 'lines',
                lon: cvzMinusInGalactic.map(p => p.l > 180 ? p.l - 360 : p.l),
                lat: cvzMinusInGalactic.map(p => p.b),
                name: 'CVZ -54° (Ecliptic)',
                line: {
                  color: '#00CED1',
                  width: 1.5,
                  dash: 'dash'
                },
                hoverinfo: 'name',
                showlegend: true
              });
            }

            const titleText = coordSystem === 'equatorial' 
              ? 'Astronomical Targets - Aitoff Projection (Equatorial)'
              : 'Astronomical Targets - Aitoff Projection (Galactic)';

            const layout = {
              title: {
                text: titleText,
                font: { size: 20 }
              },
              geo: {
                projection: {
                  type: 'aitoff',
                  rotation: {
                    lon: coordSystem === 'galactic' ? 0 : 0,
                    lat: 0
                  }
                },
                center: coordSystem === 'galactic' ? { lon: 0, lat: 0 } : { lon: 0, lat: 0 },
                showland: false,
                showlakes: false,
                showcountries: false,
                showcoastlines: false,
                bgcolor: '#0a0a1a',
                lonaxis: {
                  showgrid: true,
                  gridcolor: '#333',
                  gridwidth: 1,
                  range: [-180, 180],
                  dtick: 30,
                  tick0: 0,
                  showticklabels: true,
                  tickfont: {
                    color: '#ffffff',
                    size: 12
                  },
                  ticksuffix: '°',
                  title: {
                    text: coordSystem === 'equatorial' ? 'RA (degrees)' : 'l (degrees)',
                    font: {
                      color: '#ffffff',
                      size: 14
                    }
                  }
                },
                lataxis: {
                  showgrid: true,
                  gridcolor: '#333',
                  gridwidth: 1,
                  range: [-90, 90],
                  dtick: 30,
                  tick0: 0,
                  showticklabels: true,
                  tickfont: {
                    color: '#ffffff',
                    size: 12
                  },
                  ticksuffix: '°',
                  title: {
                    text: coordSystem === 'equatorial' ? 'Dec (degrees)' : 'b (degrees)',
                    font: {
                      color: '#ffffff',
                      size: 14
                    }
                  }
                }
              },
              paper_bgcolor: '#1a1a2e',
              plot_bgcolor: '#0a0a1a',
              font: {
                color: '#ffffff'
              },
              showlegend: true,
              legend: {
                bgcolor: '#2a2a3e',
                bordercolor: '#555',
                borderwidth: 1,
                x: 1,           // Right side (1 = far right, 0 = far left)
                y: 0,           // Bottom (0 = bottom, 1 = top)
                xanchor: 'right',  // Anchor to right edge
                yanchor: 'bottom', // Anchor to bottom edge
                orientation: 'v'   // Vertical orientation
              },
              margin: { t: 50, b: 50, l: 50, r: 50 }
            };

            const config = {
              responsive: true,
              displayModeBar: true,
              displaylogo: false,
              modeBarButtonsToAdd: ['resetGeo'],  // This adds a built-in reset button
              modeBarButtonsToRemove: ['lasso2d', 'select2d']
            };



            Plotly.newPlot('skyPlot', traces, layout, config);

          }, [plotKey, datasets, coordSystem]);

          return (
            <div className="fixed inset-0 bg-gradient-to-br from-slate-900 via-purple-900 to-slate-900 flex flex-col">
              {/* Header - Fixed at top */}
              <div className="flex-shrink-0 bg-black bg-opacity-30 backdrop-blur-sm border-b border-white border-opacity-10">
                <div className="px-4 py-3">
                  <div className="flex items-center justify-between">
                    <div>
                      <h1 className="text-2xl font-bold text-white">
                        🌌 Roman Visit Plotter
                      </h1>
                      <p className="text-gray-300 text-sm">
                        Upload CSV or Excel files with RA, DEC, and Intended_Purpose columns
                      </p>
                    </div>
                    
                    {/* Quick stats */}
                    {datasets.length > 0 && (
                      <div className="text-white text-sm">
                        Total targets: {datasets.reduce((sum, ds) => sum + ds.data.length, 0)}
                      </div>
                    )}
                  </div>
                </div>
              </div>

              {/* Controls - Fixed below header */}
              <div className="flex-shrink-0 bg-black bg-opacity-20 backdrop-blur-sm border-b border-white border-opacity-10">
                <div className="px-4 py-3">
                  <div className="flex flex-wrap items-center gap-4">
                    {/* File upload */}
                    <div className="flex items-center gap-2">
                      <label className="px-4 py-2 bg-green-600 hover:bg-green-700 text-white rounded-lg font-semibold cursor-pointer transition-colors duration-200">
                        📁 Upload File
                        <input
                          type="file"
                          accept=".csv,.xlsx,.xls"
                          onChange={handleFileUpload}
                          className="hidden"
                        />
                      </label>
                      {datasets.length > 0 && (
                        <button
                          onClick={clearAll}
                          className="px-4 py-2 bg-red-600 hover:bg-red-700 text-white rounded-lg font-semibold transition-colors duration-200"
                        >
                          Clear All
                        </button>
                      )}
                    </div>

                    {/* Coordinate system toggle */}
                    <div className="flex items-center gap-2 bg-white bg-opacity-5 px-4 py-2 rounded-lg">
                      <span className="text-gray-300 font-semibold text-sm">Coordinates:</span>
                      <button
                        onClick={() => setCoordSystem('equatorial')}
                        className={`px-3 py-1 rounded-lg font-semibold text-sm transition-colors duration-200 ${
                          coordSystem === 'equatorial'
                            ? 'bg-purple-600 text-white'
                            : 'bg-white bg-opacity-10 text-gray-300 hover:bg-opacity-20'
                        }`}
                      >
                        Equatorial
                      </button>
                      <button
                        onClick={() => setCoordSystem('galactic')}
                        className={`px-3 py-1 rounded-lg font-semibold text-sm transition-colors duration-200 ${
                          coordSystem === 'galactic'
                            ? 'bg-purple-600 text-white'
                            : 'bg-white bg-opacity-10 text-gray-300 hover:bg-opacity-20'
                        }`}
                      >
                        Galactic
                      </button>
                    </div>


                    {/* Loaded files indicator */}
                    {datasets.length > 0 && (
                      <div className="ml-auto">
                        <span className="text-gray-300 text-sm">
                          📊 {datasets.length} file{datasets.length > 1 ? 's' : ''} loaded
                        </span>
                      </div>
                    )}
                  </div>

                  {/* Loaded files list (expandable) */}
                  {datasets.length > 0 && (
                    <div className="mt-3 space-y-2">
                      {datasets.map((ds, idx) => (
                        <div
                          key={idx}
                          className="flex items-center justify-between bg-white bg-opacity-5 px-3 py-2 rounded text-sm"
                        >
                          <span className="text-gray-200">
                            {ds.fileName} ({ds.data.length} targets)
                          </span>
                          <button
                            onClick={() => removeDataset(idx)}
                            className="text-red-400 hover:text-red-300 font-semibold"
                          >
                            Remove
                          </button>
                        </div>
                      ))}
                    </div>
                  )}
                </div>
              </div>

              {/* Plot - Takes remaining space */}
              <div className="flex-1 overflow-hidden p-4">
                <div className="h-full w-full bg-white bg-opacity-5 backdrop-blur-lg rounded-lg shadow-2xl">
                  <div id="skyPlot" style={{ width: '100%', height: '100%' }}></div>
                </div>
              </div>
            </div>
          );
        }

        const root = ReactDOM.createRoot(document.getElementById('root'));
        root.render(<AstronomicalSkyPlotter />);
    </script>
</body>
</html>"""
    # Replace the placeholder with actual preloaded data
    html_content = html_content.replace('__PRELOADED_PLACEHOLDER__', preloaded_js)
    html_content = html_content.replace('__SUN_PLACEHOLDER__', sun_js)



    return html_content


def export_static_png(preloaded_datasets, output_png, coord_system='equatorial', sun_date=None):
    """
    Export a static PNG of the sky plot using plotly + kaleido.
    This reproduces the browser plot server-side.
    
    Requires: pip install plotly kaleido
    """
    try:
        import plotly.graph_objects as go
    except ImportError:
        raise ImportError(
            "plotly is required for PNG export. "
            "Install with: pip install plotly kaleido"
        )

    import math

    # --- Coordinate conversion helpers ---
    def equatorial_to_galactic(ra_deg, dec_deg):
        ra = math.radians(ra_deg)
        dec = math.radians(dec_deg)
        x_eq = math.cos(dec) * math.cos(ra)
        y_eq = math.cos(dec) * math.sin(ra)
        z_eq = math.sin(dec)
        # IAU rotation matrix
        T = [
            [-0.054875539726, -0.873437108010, -0.483834985808],
            [+0.494109453312, -0.444829589425, +0.746982251810],
            [-0.867666135858, -0.198076386122, +0.455983795705]
        ]
        x_gal = T[0][0]*x_eq + T[0][1]*y_eq + T[0][2]*z_eq
        y_gal = T[1][0]*x_eq + T[1][1]*y_eq + T[1][2]*z_eq
        z_gal = T[2][0]*x_eq + T[2][1]*y_eq + T[2][2]*z_eq
        b = math.degrees(math.asin(z_gal))
        l = math.degrees(math.atan2(y_gal, x_gal))
        if l < 0:
            l += 360
        return l, b

    def ecliptic_to_equatorial(lam_deg, beta_deg=0):
        obl = math.radians(23.43928)
        lam = math.radians(lam_deg)
        beta = math.radians(beta_deg)
        ra = math.atan2(
            math.sin(lam) * math.cos(obl) - math.tan(beta) * math.sin(obl),
            math.cos(lam)
        )
        dec = math.asin(
            math.sin(beta) * math.cos(obl) + math.cos(beta) * math.sin(obl) * math.sin(lam)
        )
        ra_deg = math.degrees(ra)
        if ra_deg < 0:
            ra_deg += 360
        return ra_deg, math.degrees(dec)

    def galactic_to_equatorial(l_deg, b_deg):
        """Inverse of equatorial_to_galactic."""
        l = math.radians(l_deg)
        b = math.radians(b_deg)
        x_gal = math.cos(b) * math.cos(l)
        y_gal = math.cos(b) * math.sin(l)
        z_gal = math.sin(b)
        # Inverse (transpose) of the IAU rotation matrix
        T = [
            [-0.054875539726, +0.494109453312, -0.867666135858],
            [-0.873437108010, -0.444829589425, -0.198076386122],
            [-0.483834985808, +0.746982251810, +0.455983795705]
        ]
        x_eq = T[0][0]*x_gal + T[0][1]*y_gal + T[0][2]*z_gal
        y_eq = T[1][0]*x_gal + T[1][1]*y_gal + T[1][2]*z_gal
        z_eq = T[2][0]*x_gal + T[2][1]*y_gal + T[2][2]*z_gal
        dec = math.degrees(math.asin(z_eq))
        ra = math.degrees(math.atan2(y_eq, x_eq))
        if ra < 0:
            ra += 360
        return ra, dec

    # --- Build traces ---
    fig = go.Figure()

    color_palette = [
        '#FF6B6B', '#4ECDC4', '#45B7D1', '#FFA07A', '#98D8C8',
        '#F7DC6F', '#BB8FCE', '#85C1E2', '#F8B739', '#52B788'
    ]

    # --- Star catalog ---
    star_catalog_file = os.path.join(os.path.dirname(__file__), 'Constellation_Stars_nolatex.csv')
    if os.path.isfile(star_catalog_file):
        star_catalog = json.loads(embed_csv_as_js_array(star_catalog_file))

        star_ras = []
        star_decs = []
        star_sizes = []
        star_opacities = []
        star_names = []

        for star in star_catalog:
            try:
                ra = float(star.get('RA [deg]', ''))
                dec = float(star.get('Dec [deg]', ''))
                vmag = float(star.get('V App Mag', '5'))
            except (ValueError, TypeError):
                continue

            star_ras.append(ra)
            star_decs.append(dec)
            star_sizes.append(max(2, 10 - vmag * 1.5))
            star_opacities.append(min(0.7, 2.5 / (vmag + 2)))

            name_parts = []
            if star.get('Common Name'):
                name_parts.append(star['Common Name'])
            if star.get('I'):
                name_parts.append(star['I'])
            if star.get('V App Mag'):
                name_parts.append(f"V={star['V App Mag']}")
            star_names.append('<br>'.join(name_parts) if name_parts else '')

        if star_ras:
            if coord_system == 'equatorial':
                s_lon = [ra - 360 if ra > 180 else ra for ra in star_ras]
                s_lat = star_decs
            else:
                gal = [equatorial_to_galactic(ra, dec) for ra, dec in zip(star_ras, star_decs)]
                s_lon = [l - 360 if l > 180 else l for l, b in gal]
                s_lat = [b for l, b in gal]

            fig.add_trace(go.Scattergeo(
                lon=s_lon,
                lat=s_lat,
                mode='markers',
                name='Catalog Stars',
                marker=dict(
                    size=star_sizes,
                    color='#FFFFE0',
                    opacity=star_opacities,
                    symbol='star',
                    line=dict(color='#FFFFFF', width=0.3)
                ),
                hoverinfo='text',
                text=star_names
            ))
            print(f"  ⭐ Added {len(star_ras)} catalog stars")
    else:
        print(f"  ⚠️  Star catalog not found at '{star_catalog_file}', skipping")

    # Process all preloaded datasets
    all_data = []
    for ds in preloaded_datasets:
        rows = json.loads(ds['data_json'])
        for row in rows:
            ra_val = row.get('RA') or row.get('ra') or row.get('Ra')
            dec_val = row.get('DEC') or row.get('dec') or row.get('Dec')
            if ra_val is None or dec_val is None:
                continue
            try:
                ra = float(ra_val)
                dec = float(dec_val)
            except (ValueError, TypeError):
                continue
            purpose = (row.get('Intended_Purpose') or row.get('intended_purpose')
                       or row.get('INTENDED_PURPOSE') or 'Unknown')
            name = row.get('Name') or row.get('name') or row.get('NAME') or ''
            all_data.append({'ra': ra, 'dec': dec, 'purpose': purpose, 'name': name})

    # Group by purpose and plot
    purposes = list(dict.fromkeys(d['purpose'] for d in all_data))  # unique, preserving order
    purpose_colors = {p: color_palette[i % len(color_palette)] for i, p in enumerate(purposes)}

    for purpose in purposes:
        pts = [d for d in all_data if d['purpose'] == purpose]
        if coord_system == 'equatorial':
            lons = [d['ra'] - 360 if d['ra'] > 180 else d['ra'] for d in pts]
            lats = [d['dec'] for d in pts]
        else:
            coords = [equatorial_to_galactic(d['ra'], d['dec']) for d in pts]
            lons = [l - 360 if l > 180 else l for l, b in coords]
            lats = [b for l, b in coords]

        fig.add_trace(go.Scattergeo(
            lon=lons, lat=lats,
            mode='markers',
            name=purpose,
            marker=dict(size=6, color=purpose_colors[purpose], opacity=0.7,
                        line=dict(color='white', width=0.5)),
            hoverinfo='text',
            text=[f"{purpose}<br>{d['name']}" if d['name'] else purpose for d in pts]
        ))

    # --- Reference lines ---
    # Ecliptic plane
    ecl_ra_dec = [ecliptic_to_equatorial(lam) for lam in range(361)]
    if coord_system == 'equatorial':
        ecl_lon = [ra - 360 if ra > 180 else ra for ra, dec in ecl_ra_dec]
        ecl_lat = [dec for ra, dec in ecl_ra_dec]
    else:
        ecl_gal = [equatorial_to_galactic(ra, dec) for ra, dec in ecl_ra_dec]
        ecl_lon = [l - 360 if l > 180 else l for l, b in ecl_gal]
        ecl_lat = [b for l, b in ecl_gal]
    fig.add_trace(go.Scattergeo(
        lon=ecl_lon, lat=ecl_lat, mode='lines',
        name='Ecliptic Plane',
        line=dict(color='#FFD700', width=2, dash='dot'),
        hoverinfo='name'
    ))

    # CVZ at +/-54 deg ecliptic latitude
    for cvz_lat, cvz_name in [(54, 'CVZ +54°'), (-54, 'CVZ -54°')]:
        cvz_ra_dec = [ecliptic_to_equatorial(lam, cvz_lat) for lam in range(361)]
        if coord_system == 'equatorial':
            cvz_lon = [ra - 360 if ra > 180 else ra for ra, dec in cvz_ra_dec]
            cvz_latvals = [dec for ra, dec in cvz_ra_dec]
        else:
            cvz_gal = [equatorial_to_galactic(ra, dec) for ra, dec in cvz_ra_dec]
            cvz_lon = [l - 360 if l > 180 else l for l, b in cvz_gal]
            cvz_latvals = [b for l, b in cvz_gal]
        fig.add_trace(go.Scattergeo(
            lon=cvz_lon, lat=cvz_latvals, mode='lines',
            name=cvz_name,
            line=dict(color='#00CED1', width=1.5, dash='dash'),
            hoverinfo='name'
        ))

    # Galactic plane
    if coord_system == 'equatorial':
        gal_pts = [galactic_to_equatorial(l, 0) for l in range(361)]
        gal_lon = [ra - 360 if ra > 180 else ra for ra, dec in gal_pts]
        gal_lat = [dec for ra, dec in gal_pts]
    else:
        gal_lon = [l - 360 if l > 180 else l for l in range(361)]
        gal_lat = [0] * 361
    fig.add_trace(go.Scattergeo(
        lon=gal_lon, lat=gal_lat, mode='lines',
        name='Galactic Plane',
        line=dict(color='#FF69B4', width=2, dash='dash'),
        hoverinfo='name'
    ))

    # --- Layout ---
    title = ('Astronomical Targets - Aitoff Projection (Equatorial)'
             if coord_system == 'equatorial'
             else 'Astronomical Targets - Aitoff Projection (Galactic)')

    fig.update_layout(
        title=dict(text=title, font=dict(size=20, color='white')),
        geo=dict(
            projection_type='aitoff',
            showland=False, showlakes=False,
            showcountries=False, showcoastlines=False,
            bgcolor='#0a0a1a',
            lonaxis=dict(showgrid=True, gridcolor='#333', gridwidth=1,
                         range=[-180, 180], dtick=30),
            lataxis=dict(showgrid=True, gridcolor='#333', gridwidth=1,
                         range=[-90, 90], dtick=30),
        ),
        paper_bgcolor='#1a1a2e',
        plot_bgcolor='#0a0a1a',
        font=dict(color='white'),
        showlegend=True,
        legend=dict(bgcolor='#2a2a3e', bordercolor='#555', borderwidth=1,
                    x=1, y=0, xanchor='right', yanchor='bottom'),
        margin=dict(t=50, b=50, l=50, r=50),
        width=1920,
        height=1080
    )

# --- Sun and anti-Sun exclusion zones + markers ---
    sun_info = get_sun_position(sun_date)
    R = sun_info['exclusion_radius']

    # Sun exclusion circle
    if coord_system == 'equatorial':
        scir_lon = [p['ra'] - 360 if p['ra'] > 180 else p['ra'] for p in sun_info['sun_circle']]
        scir_lat = [p['dec'] for p in sun_info['sun_circle']]
    else:
        scir_lon = [p['l'] - 360 if p['l'] > 180 else p['l'] for p in sun_info['sun_circle']]
        scir_lat = [p['b'] for p in sun_info['sun_circle']]

    fig.add_trace(go.Scattergeo(
        lon=scir_lon, lat=scir_lat, mode='lines',
        name=f'Sun exclusion ({R}°)',
        line=dict(color='#FF4500', width=2, dash='dash'),
        fill='toself',
        fillcolor='rgba(180, 200, 255, 0.2)',
        hoverinfo='name'
    ))

    # Anti-Sun exclusion circle
    if coord_system == 'equatorial':
        acir_lon = [p['ra'] - 360 if p['ra'] > 180 else p['ra'] for p in sun_info['anti_sun_circle']]
        acir_lat = [p['dec'] for p in sun_info['anti_sun_circle']]
    else:
        acir_lon = [p['l'] - 360 if p['l'] > 180 else p['l'] for p in sun_info['anti_sun_circle']]
        acir_lat = [p['b'] for p in sun_info['anti_sun_circle']]

    fig.add_trace(go.Scattergeo(
        lon=acir_lon, lat=acir_lat, mode='lines',
        name=f'Anti-Sun exclusion ({R}°)',
        line=dict(color='#4169E1', width=2, dash='dash'),
        fill='toself',
        fillcolor='rgba(180, 200, 255, 0.2)',
        hoverinfo='name'
    ))

    # Sun marker
    if coord_system == 'equatorial':
        s_lon = sun_info['ra'] - 360 if sun_info['ra'] > 180 else sun_info['ra']
        s_lat = sun_info['dec']
    else:
        s_lon = sun_info['l'] - 360 if sun_info['l'] > 180 else sun_info['l']
        s_lat = sun_info['b']

    fig.add_trace(go.Scattergeo(
        lon=[s_lon], lat=[s_lat], mode='markers+text',
        name=f'Sun ({sun_info["date_str"]})',
        text=['☀'], textposition='top center',
        textfont=dict(size=18, color='#FFD700'),
        marker=dict(size=14, color='#FFD700', symbol='circle',
                    line=dict(color='#FF8C00', width=2)),
        hoverinfo='text',
        hovertext=[f'Sun - {sun_info["date_str"]}']
    ))

    # Anti-Sun marker
    if coord_system == 'equatorial':
        a_lon = sun_info['anti_ra'] - 360 if sun_info['anti_ra'] > 180 else sun_info['anti_ra']
        a_lat = sun_info['anti_dec']
    else:
        a_lon = sun_info['anti_l'] - 360 if sun_info['anti_l'] > 180 else sun_info['anti_l']
        a_lat = sun_info['anti_b']

    fig.add_trace(go.Scattergeo(
        lon=[a_lon], lat=[a_lat], mode='markers',
        name=f'Anti-Sun ({sun_info["date_str"]})',
        marker=dict(size=12, color='#4169E1', symbol='circle',
                    line=dict(color='#1E3A8A', width=2)),
        hoverinfo='text',
        hovertext=[f'Anti-Sun - {sun_info["date_str"]}']
    ))

    fig.write_image(output_png, scale=2)
    print(f"  📷 Exported PNG: '{output_png}' (3840×2160)")

def main():
    examples = """\
examples:
  # Generate with no pre-loaded data (original behavior):
  %(prog)s

  # Pre-load a single CSV file:
  %(prog)s my_targets.csv

  # Pre-load an XLSX file (reads first/active sheet by default):
  %(prog)s my_targets.xlsx

  # Pre-load XLSX with a specific sheet applied to all XLSX files:
  %(prog)s my_targets.xlsx -s "Sheet2"

  # Mix CSV and XLSX files freely:
  %(prog)s science_targets.csv calibrators.xlsx

  # Multiple CSV files:
  %(prog)s cycle1_targets.csv cycle2_targets.csv cycle3_targets.csv

  # Custom output filename only (no pre-loaded data):
  %(prog)s -o custom_plotter.html

  # Override the Sun position date:
  %(prog)s targets.csv --date 2027-03-15

notes:
  - All data files must contain at least RA and DEC columns.
  - An optional Intended_Purpose column controls color-coding by category.
  - Pre-loaded data appears immediately when the HTML is opened; additional
    files can still be uploaded interactively in the browser.
  - The colon syntax (file.xlsx:SheetName) overrides the global -s/--sheet
    option for that specific file. If a file path itself contains a colon
    (e.g., C:\\data.xlsx on Windows), use the -s flag instead.
  - XLSX support requires the openpyxl package: pip install openpyxl
"""

    parser = argparse.ArgumentParser(
        description='Generate a standalone Roman Sky Plotter HTML file with '
                    'optional pre-loaded CSV/XLSX data.',
        epilog=examples,
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument(
        'data_files',
        nargs='*',
        metavar='FILE',
        help='CSV or XLSX file(s) to pre-load into the plotter. '
             'Use FILE:SHEET syntax to specify a sheet for XLSX files.'
    )
    parser.add_argument(
        '-o', '--output',
        default='roman_plotter.html',
        metavar='HTML',
        help='Output HTML filename (default: roman_plotter.html)'
    )
    parser.add_argument(
        '-s', '--sheet',
        default=None,
        metavar='NAME',
        help='Default sheet name to read from all XLSX files '
             '(default: first/active sheet). Overridden by per-file '
             'colon syntax (e.g., file.xlsx:SheetName).'
    )

    parser.add_argument(
    '--png',
    default=None,
    metavar='PNG_FILE',
    help='Also export a static PNG of the sky plot (requires plotly and kaleido). '
            'Example: --png sky_plot.png'
    )

    parser.add_argument(
        '-d', '--date',
        default=None,
        metavar='YYYY-MM-DD',
        help='Date for Sun position calculation (default: today). '
             'Example: --date 2026-06-15'
    )
    args = parser.parse_args()

    print("Generating Roman Sky Plotter HTML file...")

    preloaded_datasets = []
    for data_file in args.data_files:
        # Support "filename.xlsx:SheetName" syntax for per-file sheet selection
        if ':' in data_file and not os.path.isfile(data_file):
            filepath, sheet = data_file.rsplit(':', 1)
        else:
            filepath = data_file
            sheet = args.sheet  # global default (None = first sheet)

        if not os.path.isfile(filepath):
            print(f"⚠️  Warning: '{filepath}' not found, skipping.")
            continue

        ext = os.path.splitext(filepath)[1].lower()
        if ext not in ('.csv', '.xlsx', '.xls'):
            print(f"⚠️  Warning: '{filepath}' is not a supported format "
                  f"(.csv, .xlsx, .xls), skipping.")
            continue

        try:
            data_json = embed_file_as_js_array(filepath, sheet_name=sheet)
            n_rows = len(json.loads(data_json))
            sheet_info = f", sheet='{sheet}'" if sheet else ""
            preloaded_datasets.append({
                'fileName': os.path.basename(filepath),
                'data_json': data_json
            })
            print(f"  📂 Embedding '{filepath}'{sheet_info} ({n_rows} rows)")
        except ImportError as e:
            print(f"❌ Missing dependency: {e}")
            continue
        except Exception as e:
            print(f"⚠️  Error reading '{filepath}': {e}")
            continue


    # --- Resolve Sun date ---
    sun_date = None

    # Priority 1: explicit --date flag
    if args.date:
        from datetime import datetime, timezone
        try:
            sun_date = datetime.strptime(args.date, '%Y-%m-%d').replace(tzinfo=timezone.utc)
            print(f"  ☀️  Sun position date (from --date): {args.date}")
        except ValueError:
            print(f"⚠️  Invalid date format '{args.date}', will try data file. Expected: YYYY-MM-DD")

    # Priority 2: "Start" field from first row of first data file
    if sun_date is None and preloaded_datasets:
        first_data = json.loads(preloaded_datasets[0]['data_json'])
        if first_data:
            start_val = (first_data[0].get('Start') or first_data[0].get('start')
                         or first_data[0].get('START'))
            sun_date = parse_start_date(start_val)
            if sun_date:
                print(f"  ☀️  Sun position date (from data 'Start' field): "
                      f"{sun_date.strftime('%Y-%m-%d')} (DOY {sun_date.timetuple().tm_yday})")
            elif start_val:
                print(f"  ⚠️  Could not parse 'Start' field: '{start_val}', using today")

    # Priority 3: today
    if sun_date is None:
        from datetime import datetime, timezone
        sun_date = datetime.now(timezone.utc)

    # Compute Sun position
    sun_pos = get_sun_position(sun_date)
    print(f"  ☀️  Sun RA={sun_pos['ra']:.2f}°, Dec={sun_pos['dec']:.2f}°  "
          f"(l={sun_pos['l']:.2f}°, b={sun_pos['b']:.2f}°)")
    print(f"  🌑 Anti-Sun RA={sun_pos['anti_ra']:.2f}°, Dec={sun_pos['anti_dec']:.2f}°  "
          f"(l={sun_pos['anti_l']:.2f}°, b={sun_pos['anti_b']:.2f}°)")
    print(f"  🚫 Exclusion zones: {sun_pos['exclusion_radius']}° radius")

    html = generate_html(preloaded_datasets=preloaded_datasets, sun_position=sun_pos)

    with open(args.output, 'w', encoding='utf-8') as f:
        f.write(html)

    if args.png:
        output_png = args.png
    else:
        output_png = 'roman_plotter'
    if not preloaded_datasets:
        print("⚠️  No data files provided — skipping PNG export (nothing to plot).")
    else:
        try:
            export_static_png(preloaded_datasets, output_png+'_eq.png', coord_system='equatorial', sun_date=sun_date)
            export_static_png(preloaded_datasets, output_png+'_gal.png', coord_system='galactic', sun_date=sun_date)
        except ImportError as e:
            print(f"❌ PNG export failed: {e}")
        except Exception as e:
            print(f"⚠️  PNG export failed: {e}")

    # --- Summary ---
    n_files = len(preloaded_datasets)
    n_total = sum(len(json.loads(ds['data_json'])) for ds in preloaded_datasets)

    print(f"\n✅ Success! Generated '{args.output}'")
    if preloaded_datasets:
        print(f"   Pre-loaded {n_files} dataset(s), {n_total} total rows")
    print(f"\nTo use:")
    print(f"  1. Open '{args.output}' in any web browser")
    if preloaded_datasets:
        print(f"  2. Pre-loaded data will appear immediately on the sky plot")
        print(f"  3. Upload additional CSV/Excel files interactively if needed")
    else:
        print(f"  2. Upload CSV or Excel files with RA, DEC columns")
        print(f"  3. View your astronomical targets on an interactive Aitoff projection")
    print(f"\nThe HTML file is completely standalone and works offline!")




if __name__ == "__main__":
    main()
